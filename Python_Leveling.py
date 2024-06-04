#import relevant modules
import openpyxl
import random
import sys
import time

#Function for slower typing
def print_slow(str):
    for letter in str:
        sys.stdout.write(letter)
        sys.stdout.flush()
        time.sleep(0.05)

#Loop to ask for valid class value, then load workbook and set sheet based on Class
while True: 
    x = input("What is your class? ")
    if x.capitalize() not in (["Barbarian", "Bard", "Cleric", "Druid", "Fighter", "Monk", "Paladin", "Ranger", "Rogue", "Sorcerer", "Wizard"]):
            print("\nPlease choose an available class \nBarbarian, Bard, Cleric, Druid, Fighter, Monk, Paladin, Ranger, Rogue, Sorcerer, or Wizard\n")
            continue
    else:
        break
x = x.capitalize()
path = "Class Info.xlsx"
wb_obj = openpyxl.load_workbook(path)
wb_obj.active = wb_obj[x]
sheet_obj = wb_obj.active

#Ask for Input values of current level and HP for later variables
CurrentLV = input ("What is your current level? ")
Currenthp = input("What is your current HP? ")

#set variables for new levels
NewLV = (sheet_obj.cell(row = (2+int(CurrentLV)), column = 1).value)
NewBAB = (sheet_obj.cell(row = (2+int(CurrentLV)), column = 2).value)
NewFort = (sheet_obj.cell(row = (2+int(CurrentLV)), column = 3).value)
NewReflex = (sheet_obj.cell(row = (2+int(CurrentLV)), column = 4).value)
NewWill = (sheet_obj.cell(row = (2+int(CurrentLV)), column = 5).value)
Special1 = (sheet_obj.cell(row = (2+int(CurrentLV)), column = 6).value)
Special2 = (sheet_obj.cell(row = (2+int(CurrentLV)), column = 7).value)
Special3 = (sheet_obj.cell(row = (2+int(CurrentLV)), column = 8).value)
Special4 = (sheet_obj.cell(row = (2+int(CurrentLV)), column = 9).value)
Special5 = (sheet_obj.cell(row = (2+int(CurrentLV)), column = 10).value)
Special6 = (sheet_obj.cell(row = (2+int(CurrentLV)), column = 11).value)
Special7 = (sheet_obj.cell(row = (2+int(CurrentLV)), column = 12).value)
HD = int(sheet_obj.cell(row = 2, column = 14).value)
SkillRanks = (sheet_obj.cell(row = 2, column = 15).value)
HPRoll = (random.randrange (1,HD))

#defining function for feats
def feat():
    while True:
        Feat = input("\nPlease choose one of the following feats \n \nDodge \nCombat Reflexes \nCombat Expertise \nWeapon Expertise \n \n")
        if Feat.title() not in (["Dodge", "Combat Reflexes", "Combat Expertise", "Weapon Expertise"]):
            continue
        else:
            Feat = Feat.title()
            return Feat
            break

#defining function for special abilites
def Specials(specialNum):
    try:
        if len(specialNum) > 1:
            print(specialNum)
    except TypeError:
        return()

#Feat for odd levels
if int((NewLV)%2):
    Feat = feat()

#Output new values for character sheet
print("\n")
print_slow("Below are your new stats for leveling up! \n")
print_slow("Your level = "), print(NewLV)
if NewLV == 1:
    print_slow("New HP = "), print(HD)
else:
    print_slow("Dice Roll = "), print(HPRoll)
    NewHP = int(Currenthp) + HPRoll
    print_slow("New HP = "), print(NewHP)
print_slow("Skill ranks = "), print(SkillRanks)
print_slow("BAB = "), print(NewBAB)
print_slow("Fort Save = "), print(NewFort)
print_slow("Reflex Save = "), print(NewReflex)
print_slow("Will Save = "), print(NewWill)
print('\n')

#Run function to check for new class features
print_slow("Below are your new class features! \n")
Specials(Special1)
time.sleep(0.5)
Specials(Special2)
time.sleep(0.5)
Specials(Special3)
time.sleep(0.5)
Specials(Special4)
time.sleep(0.5)
Specials(Special5)
time.sleep(0.5)
Specials(Special6)
time.sleep(0.5)
Specials(Special7)

#Check if new feat is requied, if yes run Feat function
if (int(NewLV)%2):  
    print('\n')
    print_slow("Your New Feat = "), print_slow(Feat)
else:
    print('\n')
    print_slow("No new feat at even levels")
input('\n' 
'\n' 
"Press any key to exit")


"""Work on creating GUI"""
"""Work on creating a webscraper to find python informatoin from d20pfsrd"""